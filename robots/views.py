from django.shortcuts import render
import json
from django.http import JsonResponse
from django.views import View
from .models import Robot
from django.utils.dateparse import parse_datetime
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import HttpResponse
import openpyxl
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count

@method_decorator(csrf_exempt, name='dispatch')
class RobotCreateView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            # Валидация обязательных полей
            model = data.get("model")
            version = data.get("version")
            created = data.get("created")

            if not all([model, version, created]):
                return JsonResponse({"error": "Все поля (model, version, created) обязательны."}, status=400)

            # Проверка формата даты
            created_datetime = parse_datetime(created)
            if not created_datetime:
                return JsonResponse({"error": "Некорректный формат даты."}, status=400)

            # Сохранение робота в базу данных
            Robot.objects.create(
                serial=f"{model}-{version}",
                model=model,
                version=version,
                created=created_datetime
            )
            return JsonResponse({"message": "Робот успешно создан."}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Некорректный JSON."}, status=400)

class RobotReportView(View):
    def get(self, request, *args, **kwargs):
        # 1. Получаем данные за последнюю неделю
        last_week = timezone.now() - timedelta(days=7)

        # 2. Используем Django ORM для группировки данных по модели и версии
        robots = Robot.objects.filter(created__gte=last_week).values('model', 'version').annotate(count=Count('id'))

        # 3. Группируем данные
        report_data = {}
        for robot in robots:
            model = robot['model']
            version = robot['version']
            count = robot['count']
            if model not in report_data:
                report_data[model] = {}
            report_data[model][version] = count

        # 4. Создаем Excel-файл
        workbook = openpyxl.Workbook()

        # Дефолтный лист "Summary"
        default_sheet = workbook.active
        default_sheet.title = "Summary"
        default_sheet.append(["Модель", "Версия", "Количество за неделю"])

        # Заполнение "Summary" данными
        for model, versions in report_data.items():
            for version, count in versions.items():
                default_sheet.append([model, version, count])

        # Добавляем отдельные листы для каждой модели
        for model, versions in report_data.items():
            sheet = workbook.create_sheet(title=model)
            sheet.append(["Модель", "Версия", "Количество за неделю"])
            for version, count in versions.items():
                sheet.append([model, version, count])

            # Настройка ширины столбцов для читаемости
            for col in range(1, 4):
                column = openpyxl.utils.get_column_letter(col)
                max_length = 0
                for row in sheet.iter_rows():
                    for cell in row:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        # 5. Отдаем файл в виде HTTP-ответа
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=robot_report.xlsx'
        workbook.save(response)
        return response